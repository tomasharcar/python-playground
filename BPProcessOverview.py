#Create excel overview from the Blue Prism .xml exports
#It will collect all used VBOs and their actions, pages in the process and exceptions


def GetXMLFile():
    import easygui
    import os
    global FileName
    FileName = easygui.fileopenbox('Pick XML process export(s)', 'ProcessXML',filetypes= "*.xml", multiple=True)
    global FilePath
    for File in FileName:
        FilePath = os.path.dirname(File)
        if File.find('.xml')==-1:
            print('Not XML!')
            break
    return FileName

def ReadXMLFile():
    import xml.etree.ElementTree as ET
    MultipleProcess = []
    
    for File in GetXMLFile():
        ProcessCollection ={}
        tree = ET.parse(File)
        root = tree.getroot()
        ProcessCollection['ProcessName'] = root.attrib['name']
        VBOs = []

        for obj in root.iter('resource'):
            VBOs.append(obj.attrib['object'])
        VBOs = list(set(VBOs))
        VBOs.sort()
        Actions = []
        FinalColl = {}
        for VBO in VBOs:
            for stage in root.iter('resource'):
                if stage.attrib['object']==VBO:
                    Actions.append(stage.attrib['action'])
            Actions = list(set(Actions))
            Actions.sort()
            FinalColl[VBO]=Actions
            Actions=[]
        ProcessCollection['VBOs'] = FinalColl
        FinalColl=[]
        for obj in root.iter('exception'):
            EXCs={}
            EXCs['Exception']=obj.attrib['type']
            EXCs['Detail']=obj.attrib['detail']
            if EXCs['Exception']!='' and EXCs['Detail']!='':FinalColl.append(EXCs)
        ProcessCollection['EXCs'] = FinalColl
        PGE = []
        for pge in root.findall('subsheet'):
	        PGE.append(pge.find('name').text)
        ProcessCollection['PGEs'] = PGE
        MultipleProcess.append(ProcessCollection)
    return MultipleProcess

def CreateExcelOverview(Collection):
    import openpyxl
    from openpyxl.styles import Font
    for Process in Collection:
        wb = openpyxl.Workbook()
        sheet = wb['Sheet']
        sheet.title='Pages'
        sheet['A1'] = 'Name'
        sheet['A1'].font = Font(bold=True)
        r=2
        for object in Process.get('PGEs'):
            sheet.cell(row=r, column=1).value = object
            r=r+1

        sheet = wb.create_sheet('VBOs')
        sheet['A1'] = 'Name'
        sheet['B1'] = 'Type'
        sheet['C1'] = 'Action'
        sheet['A1'].font = Font(bold=True)
        sheet['B1'].font = Font(bold=True)
        sheet['C1'].font = Font(bold=True)
        r=2

        for object in Process.get('VBOs').keys():
            sheet.cell(row=r, column=1).value = object
            sheet.cell(row=r, column=2).value = 'VBO'
            r=r+1
        r=2
        for action in Process.get('VBOs').values():
            sheet.cell(row=r, column=3).value = '; '.join(action)
            r=r+1

        sheet = wb.create_sheet('Exceptions')
        sheet['A1'] = 'Exception Type'
        sheet['B1'] = 'Detail'
        sheet['A1'].font = Font(bold=True)
        sheet['B1'].font = Font(bold=True)
        r=2
        for object in Process.get('EXCs'):
            sheet.cell(row=r, column=1).value = object['Exception']
            sheet.cell(row=r, column=2).value = object['Detail']
            r=r+1
        
    
        wb.save(FilePath+'\\'+str(Process.get('ProcessName'))+'_Overview.xlsx')


CreateExcelOverview(ReadXMLFile())





